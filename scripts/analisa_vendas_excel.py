import openpyxl

file_path = 'public/VENDAS DEZEMBRO 2025.xlsx'
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

# Procurar a primeira linha útil (com colunas de fato)
for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
    if row and any(row):
        print(f'Linha {i}:', row)
        if any(isinstance(cell, str) and cell.strip().upper() == 'CONSULTOR' for cell in row):
            headers = [str(cell).strip() if cell else '' for cell in row]
            header_row = i
            break
else:
    raise Exception('Cabeçalho não encontrado')

# Ler os dados a partir da linha do cabeçalho
rows = []
for row in sheet.iter_rows(min_row=header_row+1, values_only=True):
    if not any(row):
        continue
    rows.append(dict(zip(headers, row)))

print('Headers:', headers)
print('Primeiras linhas:', rows[:5])
